# script to run against files to create analytics data in BATCH rather than 1 spreadsheet at a time. Will only JOIN data, will not do the math on the numbers imported.
# script also outputs a web-friendly pie chart graphics file for every collection
# potential to illustrate the value of various collection sets, and to look pretty
import pandas as PD
import os
import matplotlib.pyplot as plt
import numpy as np

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

print("make sure data clean-up was completed for analytics spreadsheet")
print("!!!WARNING: XLSX with UUIDs should be in same directory as pythonScript!!!")
print("!!!WARNING: WILL CRAWL subdirectories, proceed accordingly!!!")
print("!!!Warning: stats file should be in subdirectory with yyyy-mm patter!!!")

# get the analytics file
seriousFilepath = input('source directory for stats files: ')
stats_filename = input("name of the stats file including filepath: ")
print("!!!Use same output directory as stats file!!!")
out_directory = input("Finished product directory: ")
# get date variables
month = input("month_name: ")
year = input("year: ")
# set counters for the LSTA v. Born digital pie charts at end of the script
LSTA = 0
BornDigital = 0
# Create lists
LstaLabels = []
LstaTitle = 'LSTA hits by collection'
LstaSizes = []
lsta_headers = ["Collections", "hits"]
lsta_workbook_name = out_directory + "/lsta_hits.xlsx"
lsta_wb = Workbook()
lsta_page = lsta_wb.active
lsta_page.title = "Sheet1"
lsta_page.append(lsta_headers)
lsta_wb.save(filename=lsta_workbook_name)

BornDigitalLabels = []
BornDigitalTitle = 'Born digital hits by collection'
BornDigitalSize = []
bornDigital_headers = ["Collections", "hits"]
bornDigital_workbook_name = out_directory + "/bornDigital_hits.xlsx"
bornDigital_wb = Workbook()
bornDigital_page = bornDigital_wb.active
bornDigital_page.title = "Sheet1"
bornDigital_page.append(bornDigital_headers)
bornDigital_wb.save(filename=bornDigital_workbook_name)
# borndigital_file = open("borndigital_file.csv", "a")
# borndigital_file.write("Collections,hits" + "\n")
# create master totals list
master_headers = ["Collections", "hits"]
master_workbook_name = out_directory + "/grand_totals_hits.xlsx"
master_wb = Workbook()
master_page = master_wb.active
master_page.title = "Sheet1"
master_page.append(master_headers)
master_wb.save(filename=master_workbook_name)
# master_hits = open("grand_totals.csv", "a")

# iterate over directory
for dirpath, dirnames, filenames in os.walk(seriousFilepath):
    for filename in filenames:
        if filename.endswith(('.xlsx')):
            filename2 = os.path.join(dirpath, filename)
            labels = []
            size = []
            # load stats file and UUIDs file
            df1 = PD.read_excel(filename2, sheet_name='Sheet1', engine="openpyxl")
            df2 = PD.read_excel(stats_filename, sheet_name='processed', engine="openpyxl")
            # make sum of the overall stats values
            gross = df2['Pageviews'].sum()
            # start pie chart making and create a title, note use of iloc[]
            # start pie chart creation
            labels.append(df1.iloc[0, 0])
            labels.append("Other")
            title = 'Hits for ' + df1.iloc[0, 0] + ' for ' + month + ', ' + year + "/" + str(gross) + ' total'
            # run stats inner join and dedupe results if needed
            results = df1.merge(df2, on='UUIDs', how="inner")
            results = results.drop_duplicates()
            tempMath = results['Pageviews'].sum()
            collectionTitle = df1.iloc[0, 0]
            numbersList = str(tempMath)
            # assign the sum of the results to either LSTA or born digital collection types
            if df1.iloc[0, 1] == 'LSTA':
                LSTA = LSTA + tempMath
                LstaLabels.append(collectionTitle)
                LstaSizes.append(tempMath)
                lsta_page.append([collectionTitle, numbersList])
                lsta_wb.save(lsta_workbook_name)
            # lsta_file.write(collectionTitle + "," + numbersList + "\n")
            if df1.iloc[0, 1] == 'Born-digital':
                BornDigital = BornDigital + tempMath
                BornDigitalLabels.append(collectionTitle)
                BornDigitalSize.append(tempMath)
                bornDigital_page.append([collectionTitle, numbersList])
                bornDigital_wb.save(bornDigital_workbook_name)
            # borndigital_file.write(collectionTitle + "," + numbersList + "\n")
            master_page.append([collectionTitle, numbersList])
            master_wb.save(master_workbook_name)
            # master_hits.write(collectionTitle + "," + numbersList + "\n")

            # get sum of hits for this collection spreadsheet after the merger
            # make first pie slice
            legendLabels = []
            collectionNumbers = df1.iloc[0, 0] + "\n" + str(tempMath) + ' Hits'
            legendLabels.append(collectionNumbers)
            other = gross - tempMath
            otherNumbers = "Other" + "\n" + str(other) + ' Hits'
            legendLabels.append(otherNumbers)
            sizes = [tempMath, other]
            explode = (0.1, 0)
            # construct the chart
            fig1, ax1 = plt.subplots(figsize=(6, 6))
            wedges, text, autotexts = ax1.pie(sizes, explode=explode, labels=labels, shadow=True, startangle=90,
                                              autopct='%1.1f%%', textprops=dict(color="w", weight="bold"))
            ax1.axis('equal')
            ax1.legend(legendLabels, title='Collections', loc="lower right")
            ax1.set_title(title)
            # write stats results to a file
            writer = PD.ExcelWriter(out_directory + "/" + filename, engine='xlsxwriter')
            results.to_excel(writer, "Sheet1", index=False)
            writer.save()
            pieFile = filename[:-4] + 'png'
            fig1.savefig(out_directory + "/" + pieFile, format='png')
            # write pie chart to a web-friendly svg file
            print(filename + " is processed")
            plt.close()
# close all master lists
lsta_wb.close()
bornDigital_wb.close()
master_wb.close()
# create count of queries and static page views
queries = df2[df2['UUIDs'].str.contains("/?s=")]
queries_count = queries['Pageviews'].sum()
pages = df2[df2['UUIDs'].str.contains("/")]
pages = pages[~pages['UUIDs'].str.contains("/?s=")]
pages_count = pages['Pageviews'].sum()
# start summary chart
labels = []
labels.append("LSTA")
labels.append("BornDigital")
labels.append("Queries")
labels.append("Pages")
labels.append("Other")
title = "Hits for " + month + ", " + year + "/" + str(gross) + " total"
sizes = []
sizes.append(LSTA)
sizes.append(BornDigital)
sizes.append(queries_count)
sizes.append(pages_count)
other = LSTA + BornDigital + queries_count + pages_count
other = gross - other
sizes.append(other)
explode = (0.1, 0.1, 0, 0, 0)
fig1, ax1 = plt.subplots(figsize=(6, 6))
wedges, text, autotexts = ax1.pie(sizes, explode=explode, labels=labels, shadow=True, startangle=90, autopct='%1.1f%%',
                                  textprops=dict(color="w", weight="bold"))
ax1.axis('equal')
ax1.legend(labels, title="Major types", loc="lower right")
ax1.set_title(title)
summaryGraph = out_directory + "/summaryGraph.png"
fig1.savefig(summaryGraph, format="png")
# clear labels and sizes
labels = []
sizes = []
# start top 5 chart for lsta
df1 = PD.read_excel(lsta_workbook_name, engine="openpyxl")
df1 = df1.sort_values(by=['hits'], ascending=False)
writer = PD.ExcelWriter(out_directory + "/" + "lsta_summary.xlsx", engine='xlsxwriter')
df1.to_excel(writer, "Sheet1", index=False)
writer.save()
title = "Top 5 hits for LSTA collections over total for " + month + ", " + year
labels.append(df1.iloc[0, 0] + "(" + str(df1.iloc[0, 1]) + " hits)")
labels.append(df1.iloc[1, 0] + "(" + str(df1.iloc[1, 1]) + " hits)")
labels.append(df1.iloc[2, 0] + "(" + str(df1.iloc[2, 1]) + " hits)")
labels.append(df1.iloc[3, 0] + "(" + str(df1.iloc[3, 1]) + " hits)")
labels.append(df1.iloc[4, 0] + "(" + str(df1.iloc[4, 1]) + " hits)")
sizes.append(df1.iloc[0, 1])
sizes.append(df1.iloc[1, 1])
sizes.append(df1.iloc[2, 1])
sizes.append(df1.iloc[3, 1])
sizes.append(df1.iloc[4, 1])
lsta_other = LSTA - df1.iloc[0, 1] - df1.iloc[1, 1] - df1.iloc[2, 1] - df1.iloc[3, 1] - df1.iloc[4, 1]
labels.append("All others (" + str(lsta_other) + " hits)")
sizes.append(lsta_other)
fig1, ax1 = plt.subplots(figsize=(6, 6))
wedges, text, autotexts = ax1.pie(sizes, labels=labels, shadow=True, startangle=90, autopct='%1.1f%%',
                                  textprops=dict(color="w", weight="bold"))
ax1.axis('equal')
ax1.legend(labels, title="LSTA top 5", loc="lower right")
ax1.set_title(title)
summaryGraph = out_directory + "/summary_lsta_top5.png"
fig1.savefig(summaryGraph, format="png")
# clear labels and sizes
labels = []
sizes = []
# start top 5 chart for born digital
df1 = PD.read_excel(bornDigital_workbook_name, engine="openpyxl")
df1 = df1.sort_values(by=['hits'], ascending=False)
writer = PD.ExcelWriter(out_directory + "/" + "born_digital_summary.xlsx", engine='xlsxwriter')
df1.to_excel(writer, "Sheet1", index=False)
writer.save()
title = "Top 5 hits for Born digital collections over total for " + month + ", " + year
labels.append(df1.iloc[0, 0] + "(" + str(df1.iloc[0, 1]) + " hits)")
labels.append(df1.iloc[1, 0] + "(" + str(df1.iloc[1, 1]) + " hits)")
labels.append(df1.iloc[2, 0] + "(" + str(df1.iloc[2, 1]) + " hits)")
labels.append(df1.iloc[3, 0] + "(" + str(df1.iloc[3, 1]) + " hits)")
labels.append(df1.iloc[4, 0] + "(" + str(df1.iloc[4, 1]) + " hits)")
sizes.append(df1.iloc[0, 1])
sizes.append(df1.iloc[1, 1])
sizes.append(df1.iloc[2, 1])
sizes.append(df1.iloc[3, 1])
sizes.append(df1.iloc[4, 1])
born_digital_other = BornDigital - df1.iloc[0, 1] - df1.iloc[1, 1] - df1.iloc[2, 1] - df1.iloc[3, 1] - df1.iloc[4, 1]
labels.append("All others (" + str(born_digital_other) + " hits)")
sizes.append(born_digital_other)
fig1, ax1 = plt.subplots(figsize=(6, 6))
wedges, text, autotexts = ax1.pie(sizes, labels=labels, shadow=True, startangle=90, autopct='%1.1f%%',
                                  textprops=dict(color="w", weight="bold"))
ax1.axis('equal')
ax1.legend(labels, title="Born Digital top 5", loc="lower right")
ax1.set_title(title)
summaryGraph = out_directory + "/summary_bd_top5.png"
fig1.savefig(summaryGraph, format="png")
# clear labels and sizes
labels = []
sizes = []
# start top 5 chart for grand totals
df1 = PD.read_excel(master_workbook_name, engine="openpyxl")
df1 = df1.sort_values(by=['hits'], ascending=False)
writer = PD.ExcelWriter(out_directory + "/" + "grand_totals.xlsx", engine='xlsxwriter')
df1.to_excel(writer, "Sheet1", index=False)
writer.save()
tempMath = df1['hits'].sum()
title = "Top 5 overall collections over total for " + month + ", " + year
labels.append(df1.iloc[0, 0] + "(" + str(df1.iloc[0, 1]) + " hits)")
labels.append(df1.iloc[1, 0] + "(" + str(df1.iloc[1, 1]) + " hits)")
labels.append(df1.iloc[2, 0] + "(" + str(df1.iloc[2, 1]) + " hits)")
labels.append(df1.iloc[3, 0] + "(" + str(df1.iloc[3, 1]) + " hits)")
labels.append(df1.iloc[4, 0] + "(" + str(df1.iloc[4, 1]) + " hits)")
sizes.append(df1.iloc[0, 1])
sizes.append(df1.iloc[1, 1])
sizes.append(df1.iloc[2, 1])
sizes.append(df1.iloc[3, 1])
sizes.append(df1.iloc[4, 1])
fig1, ax1 = plt.subplots(figsize=(6, 6))
wedges, text, autotexts = ax1.pie(sizes, labels=labels, shadow=True, startangle=90, autopct='%1.1f%%',
                                  textprops=dict(color="w", weight="bold"))
ax1.axis('equal')
ax1.legend(labels, title="Overall top 5", loc="lower right")
ax1.set_title(title)
summaryGraph = out_directory + "/summary_top5.png"
fig1.savefig(summaryGraph, format="png")
print("all done with stats and charts")